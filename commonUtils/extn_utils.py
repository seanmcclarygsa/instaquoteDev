import shutil
import os
import logging
#from tktooltip import ToolTip
import myConstants as mc
import numpy as np
import pyodbc as po
import zipfile
from datetime import datetime
import openpyxl as op
import glob
import platform
import pandas as pd
import sqlalchemy as sa
from sqlalchemy import create_engine, text #added by Satish 12/21/23

import google.oauth2.service_account
import googleapiclient.discovery

#fileConfig('logging_conf.ini', disable_existing_loggers=False)
#log = logging.getLogger(__name__)
#DEBUG_LEVEL = 10 # logging debug level from ini config file
FGCOLOR=mc.FGCOLOR
RESET=mc.RESET

def findMatchingRowFromDataFrame(df, paramName, paramValue):
    return df.loc[df[paramName] == paramValue]

def getDTS():
    dts = datetime.fromtimestamp(datetime.now().timestamp()).strftime("%m%d%Y_%H%M%S")
    return dts

def cleanStringColumns(dataframe, stringColumns):
    for x in stringColumns:
        dataframe[x] = dataframe[x].apply(lambda _: str(_)) #Make specified columns as strings
        dataframe[x] = dataframe[x].str.strip() #Ltrim and rtrim the string columns
#    return dataframe

def cleanIntegerColumns(dataframe, integerColumns):
    for x in integerColumns:
        dataframe[x] = dataframe[x].apply(lambda _: int(_)) #Make specified columns as integers
#    return dataframe

def cleanDecimalColumns(dataframe, decimalColumns):
    for x in decimalColumns:
        dataframe[x] = dataframe[x].apply(lambda _: round(_,2)) #Round specified columns to 2 decimals
#    return dataframe

def zipFromSourceFolder(src_dir, archive_dest, file_ext):
    files = os.listdir(src_dir) # getting all the files in the source directory
#    archive_dest = src_dir+"\\output.zip"
    for fname in files:
        name, ext = os.path.splitext(fname)
#        ext = ext.replace(".","")
        file_path = os.path.join(src_dir, fname)
        if file_ext is None:
            with zipfile.ZipFile(archive_dest, 'a',
                     compression=zipfile.ZIP_DEFLATED,
                     compresslevel=9) as zf:
                zf.write(file_path, os.path.relpath(file_path, src_dir))
        elif file_ext == ext:
            with zipfile.ZipFile(archive_dest, 'a',
                     compression=zipfile.ZIP_DEFLATED,
                     compresslevel=9) as zf:
                zf.write(file_path, os.path.relpath(file_path, src_dir))
    logging.info(f"{FGCOLOR}Successfully zipped to a file within {src_dir} ...")

# copying all the files from source directory to the destination directory
def copyFromSourceFolderToDestFolder(src_dir, dest_dir, file_ext=None): 
    files = os.listdir(src_dir) # getting all the files in the source directory
    for fname in files:
        name, ext = os.path.splitext(fname)
#        ext = ext.replace(".","")
        file_path = os.path.join(src_dir, fname)
        if file_ext is None:
           shutil.copy2(file_path, dest_dir) 
#           print("Successfully copied "+ fname+ " to the SQL Server")
           logging.info(f"{FGCOLOR}Successfully copied {fname}{RESET} to the SQL Server - from {src_dir} to {dest_dir} ...")
        elif file_ext.replace(".","") == ext.replace(".",""):
           shutil.copy2(file_path, dest_dir)
           logging.info(f"{FGCOLOR}Successfully copied {fname}{RESET} to the SQL Server - from {src_dir} to {dest_dir} ...")

# copying all the files from source directory to the destination directory
def copySelectedFilesFromSourceFolderToDestFolder(src_dir, dest_dir, files, file_ext=None): 
    for fname in files:
        name, ext = os.path.splitext(fname)
#        ext = ext.replace(".","")
        file_path = os.path.join(src_dir, fname)
        if file_ext is None:
           shutil.copy2(file_path, dest_dir) 
#           print("Successfully copied "+ fname+ " to the SQL Server")
           logging.info(f"{FGCOLOR}Successfully copied {fname}{RESET} to the SQL Server - from {src_dir} to {dest_dir} ...")
        elif file_ext.replace(".","") == ext.replace(".",""):
           shutil.copy2(file_path, dest_dir)
           logging.info(f"{FGCOLOR}Successfully copied {fname}{RESET} to the SQL Server - from {src_dir} to {dest_dir} ...")       
   

#Method to delete all contents in a given directory, or contents with a particular extension from a given directory
def deleteFolderContents(directory, file_ext=None):
    files = os.listdir(directory) # getting all the files in the source directory
#    print (files)
    for fname in files:
        name, ext = os.path.splitext(fname)
        file_path = os.path.join(directory, fname)
#        print(file_path)
        try:
            if file_ext is None:
                rmfiles(file_path)
            elif file_ext == ext:
                rmfiles(file_path)
        except Exception as e:
            logging.info(f"Failed to delete {file_path}{RESET}. Reason: {e}")
    logging.info(f"{FGCOLOR}Successfully removed files from the {directory} directory ...{RESET}")

def rmfiles(file_path):
    if os.path.isfile(file_path) or os.path.islink(file_path):
        os.remove(file_path)
    elif os.path.isdir(file_path):
        shutil.rmtree(file_path)

def dropEmptyRowsforColumns(dataframe, cols):
    nan_value = float("NaN")
    for column in dataframe[cols]:
#        val = dataframe[column]
        dataframe[column].replace(r'\s+', "", regex=True, inplace=True)
        dataframe[column].replace("", nan_value, inplace=True)
#        dataframe[column].replace("n/a", nan_value, inplace=True)
#        dataframe[column].replace("N/A", nan_value, inplace=True)
#    dataframe[cols]=dataframe[cols].replace(r'\s+', '', regex=True)
#    dataframe.replace("", nan_value, inplace=True)
    dataframe.replace("n/a", nan_value, inplace=True)
    dataframe.replace("N/A", nan_value, inplace=True)
    dataframe.dropna(subset=cols, inplace=True)
        
def setLoggingLevelForModules(moduleList, logLevel):
    for module in moduleList:
        logging.getLogger(module).setLevel(logLevel)
        
'''def createToolTip(widget, message):
    return ToolTip(widget, msg=message,delay=0,
       parent_kwargs={"bg": "black", "padx": 5, "pady": 5},
       fg="#ffffff", bg="#1c1c1c", padx=10, pady=10)'''

def connectSQL(database):
    # Connection string
    cnxn = po.connect("Driver={SQL Server Native Client 11.0};"
                          "Server=E04TCM-GSASQL03;"
                          "Database=" + database +";"
                          "Trusted_Connection=yes;") 
 
    return cnxn

def executequery(query, url):
    #url = "mssql+pyodbc://BODDBServer"
    engine = create_engine(url)
    try:
       dataframe = pd.read_sql(query, engine)
    except Exception as e:
        print("An error occured", str(e))
        print("Trying use to connect instead..")
        query = text(query)
        conn= engine.connect()
        dataframe = pd.read_sql_query(query, conn)
        conn.close()
    return dataframe

def truncateTable(database,table):
    try:
        cnxn = connectSQL(database)
        cursor = cnxn.cursor()
        if cursor is not None:
           print("database connected.")
           truncateSql = f"TRUNCATE TABLE {table}"
           cursor.execute(truncateSql)
           cnxn.commit()
        # Close the cursor and connection
        cursor.close()
        cnxn.close()
    except Exception as e:
        return print("Error: %s" % e)    
    
def execute_SP(database, sp, params=None):
    #make sure the sp has SET NOCOUNT ON at the top.
    try:
        cnxn = connectSQL(database)
        cursor = cnxn.cursor()
        if cursor is not None:
           print("database connected.")
           if params is not None:
            #construct the string for all the parameters
                paramString=''
#                database='QS_QUERY'
#                sp='sp_test'
#                for x in params:
#                    paramString += x + ' ' 
                paramString = ', '.join(params)
                query= f"Exec [{database}].dbo.[{sp}] {paramString}"
                print(query)
                cursor.execute(query)   
           else:
                query= f"Exec [{database}].dbo.[{sp}]"
                print(query)
                cursor.execute(query)
           '''
           row = cursor.fetchone()
           while row:
            # Print the row
                print(row)
                row = cursor.fetchone() # retrieves the next row of a query result set
             # Close the cursor and delete it
           '''
           rows = cursor.fetchall()
           row_string = "\n".join(map(str, rows))

        cursor.close()
        cnxn.commit()
        del cursor
        # Close the database connection
        cnxn.close() 
        print('cursor closed.')
        return row_string
       
    except Exception as e:
        return print("Error: %s" % e)

def deleteFilesFromOutputFolder(foldername):
   files = glob.glob(foldername)
   try:
      for f in files:
        os.remove(f)
   except Exception as e:
        print('Failed to delete', (e))
        
def getEmailBodyFromHTMLFile(path):
    # Opening the html file
    #HTMLFile = open(path,, errors='ignore', "r")
  
    # Reading the file
   # index = HTMLFile.read()
    with open(path, "r", encoding="utf-8") as f:
        index = f.read()
    return index 

def downloadFileFromGoogleDrive(file, destFolder, cred):
	import io
	from googleapiclient.http import MediaIoBaseDownload
	drive_service = DriveService(cred)
	file_id = file.get('id')
	fileName = file.get('name')
	if destFolder is not None:
		fileName = destFolder+fileName
	request = drive_service.files().get_media(fileId=file_id)
	fsrc = io.BytesIO()
	downloader = MediaIoBaseDownload(fsrc, request)
	done = False
	while done is False:
		status, done = downloader.next_chunk()
		print("Download %d%%" % int(status.progress() * 100))

# The file has been downloaded into RAM, now save it in a file
	fsrc.seek(0)
	with open(fileName, 'wb') as fdest:
		shutil.copyfileobj(fsrc, fdest, length=131072)

def getFilesInGoogleDriveFolder(folderID, mimeType, cred):
	drive_service = DriveService(creds=cred)
	q = "mimeType = '"+mimeType+"' and '"+folderID+"' in parents"
#	q = "mimeType='application/vnd.google-apps.folder' and 'folderID' in parents"
	# execute query
	results = drive_service.files().list(q=q).execute()
	
	# get list of files
	files = results.get('files', [])
	
	# loop through the files and print name
	for file in files:
	    #print(file['name'])
	 return files

class DriveService(object):
	_instance = None
	def __new__(cls, *args, **kwargs):
		if cls._instance is None:
			print('Creating the object')
			#	scopes = ["https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"]
#			credential_file = "auth/gsa.gov_api-project-673153394971-d38e0ee051e1.json"
			#file_id = '1TavOkhuhPusSfJ1xtma_sAgUH8Adu-z8'#this is a test
			for key, value in kwargs.items():
				if (key == "creds"):
					credential_file=value
			#credential_file = args[0]
			cls._instance = googleapiclient.discovery.build('drive', 'v3', credentials = google.oauth2.service_account.Credentials.from_service_account_file(credential_file))
			# Put any initialization here.
		return cls._instance

#	def __init__(self, cred):
#		self.credential_file = cred
#		self = googleapiclient.discovery.build('drive', 'v3', credentials = google.oauth2.service_account.Credentials.from_service_account_file(cred))

def print_colored(message, color):
    color_codes = {
        'red': '\033[91m',
        'green': '\033[92m',
        'yellow': '\033[93m',
        'blue': '\033[94m',
        'purple': '\033[95m',
        'cyan': '\033[96m',
        'white': '\033[97m',
        'reset': '\033[0m'
    }

    if color not in color_codes:
        print("Invalid color. Supported colors: red, green, yellow, blue, purple, cyan, white")
        return

    colored_message = f"{color_codes[color]}{message}{color_codes['reset']}"
    print(colored_message)

def get_os_info():
    system = platform.system()  # Get the system/OS name

    if system == "Windows":
        return "Windows"
    elif system == "Linux":
        return "Linux"
    elif system == "Darwin":
        return "macOS"
    else:
        return "Unknown"
        

def setColumnWidthDynamically(filename):
    workbook = op.load_workbook(filename)
    worksheet = workbook.active
    row_count = worksheet.max_row
    for col_num, col_name in enumerate(worksheet.iter_cols(min_row = 1, max_row=row_count)):
        max_len = max(len(str(cell.value)) for cell in col_name)
        worksheet.column_dimensions[op.utils.get_column_letter(col_num+2)].width = max_len
    workbook.save(filename)


if __name__ == '__main__':
 #   copyFromSourceFolderToDestFolder("C:\work\gsa\gss2017\pythoncode\R6automation\output","Z:\DATALOAD\Satish\MRO_PN_TSVS",".tsv")
#     deleteFolderContents("output")
#     zipFromSourceFolder("C:\\work\\gsa\\gss2017\\pythoncode\\tdr\\output","C:\\work\\gsa\\gss2017\\pythoncode\\tdr\\output\\output.zip",".tsv")
#     print(getCurrentTimeStamp())
    execute_SP('QS_QUERY','import_ADV_PICS_Session_by_Number','1150','319_04_00')
    